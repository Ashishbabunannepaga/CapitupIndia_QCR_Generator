import streamlit as st
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Safe imports from utilities to prevent circular dependencies
from utils import process_pdf_document, APIKeyManager, render_underwriter_audit_panel

HEALTH_SYSTEM_PROMPT = """
You are an expert, elite-level health insurance underwriter and auditor. Your task is to extract variables from the provided document (GMC quotation, policy schedule, or retail family floater) with absolute analytical precision.

*** STRICT UNDERWRITING RULES FOR ACCURACY ***

1. THE 18% GST BACK-CALCULATION RULE (CRITICAL): 
   - Health insurance premiums in India are subject to exactly 18% GST.
   - If the document only displays a single prominent "Total Premium Payable" (inclusive of tax) and does not explicitly itemize GST, or if your extracted GST is mathematically impossible (e.g. 0.24 or 0 on a premium in the thousands), you MUST perform a standard back-calculation:
     Net Premium (Basic) = Total Premium / 1.18
     GST (18%) = Total Premium - Net Premium (Basic)
   - Never output GST as 0.00 or a fraction if the total premium is in the thousands. Cross-verify your numbers: Net Premium + GST must equal Total Premium exactly.

2. THE ZERO-VALUE RULE: If any limit, sublimit, or benefit (e.g. Maternity C-Section, Corporate Buffer, Cataract Sublimit) is listed as "0", "0.00", "Nil", "N/A", "Not Covered", or is left blank/dash, you MUST classify it as "No" or "Not Covered".

3. GMC VS RETAIL FLOATER: Detect if the policy is a "Group Medical Cover (GMC)" or a retail "Family Floater / Individual" plan. 
   - For retail/floater plans, extract the proposer's name as the "insured_name" (e.g. Vemuri Murali Krishna).

4. ANCHORING AUDIT TRAIL: In your JSON response, you must populate the "audit_trail" object. Explain precisely the page, table, and formulas you used to cross-verify that: Net Premium + GST = Total Premium Payable.

Extract and format the output as a clean JSON object matching this exact schema:
{
  "insurer_name": "Name of insurer, e.g. STAR HEALTH, CARE HEALTH, NIVA BUPA, ICICI LOMBARD, TATA AIG",
  "insured_name": "Name of Corporate Insured / Proposer Name",
  "policy_type": "Group Medical Cover (GMC) or Family Floater or Individual",
  "employee_count": "Total covered lives or employee count, e.g. 4 Lives or 250 Lives",
  "family_definition": "Family definition, e.g. 1+3 (Self + Spouse + 2 Children) or Floater",
  "relations_covered": "Eligible relations, e.g. Self, Spouse, Children, Parents",
  "min_age_limit": "Minimum entry age, e.g. 91 Days or 18 Years",
  "max_age_limit": "Maximum entry age, e.g. 65 Years or No Limit",
  "sum_insured": "Sum Insured structure, e.g. Family Floater of 10,00,000",
  "gross_premium": 40100.0,
  "gst": 7218.0,
  "total_premium": 47318.0,
  "coverages": {
    "Room Rent Limits": "e.g. 1% of Sum Insured or Single Private AC Room",
    "ICU Limit": "e.g. 2% of Sum Insured or No Limit",
    "Pre-Existing Disease waiting period": "e.g. Covered from Day 1 / Day 1 Waiver",
    "Maternity Limit (Normal)": "e.g. 50,000 or Not Covered",
    "Maternity Limit (C-Section)": "e.g. 75,000 or Not Covered",
    "Maternity Waiting Period": "e.g. Day 1 Covered or 9 Months",
    "Pre & Post Natal Expenses": "e.g. Covered within Maternity Limit or up to 5,000",
    "New Born Baby Cover": "e.g. Covered from Day 1 within Sum Insured or up to 10% of SI",
    "Road Ambulance Limit": "e.g. up to 2,000 per hospitalization or Actuals",
    "Pre & Post Hospitalization Period": "e.g. 30 Days Pre & 60 Days Post",
    "AYUSH Treatment": "e.g. Covered up to Sum Insured or No",
    "Disease Sublimits": "e.g. No Sublimits or List of disease-specific capping",
    "Cataract Sublimit": "e.g. 30,000 per eye or No Limit",
    "Internal Congenital Disease": "e.g. Covered up to Sum Insured or No",
    "External Congenital Anomaly": "e.g. Covered up to 1,00,000 or No",
    "FESS (Sinus Surgery)": "e.g. Covered up to SI or Sublimit of 50,000",
    "Psychiatric & Mental Illness": "e.g. Covered up to Sum Insured or No",
    "Modern Treatments": "e.g. Covered up to SI or capped at 50% of SI",
    "Lasik Cover": "e.g. Covered if > 7.5 dioptres or No",
    "Day Care Treatment": "e.g. All Day Care procedures covered or No",
    "Domiciliary Hospitalization": "e.g. Covered up to Sum Insured or No",
    "Organ Donor Expenses": "e.g. Covered up to SI or No",
    "Well Mother Cover": "e.g. Covered up to 2,00,0 or No",
    "Co-payment": "e.g. No Co-pay or 10% Co-pay on parents",
    "Special Conditions": "e.g. Mid-term inclusion allowed for newborns and newly wed spouses"
  },
  "audit_trail": {
    "premium_location": "Page X, Table Y proving premium details",
    "maternity_verification": "Page X, Table Y proving maternity limits and waiting periods"
  }
}
Respond ONLY with raw JSON. Do not include markdown blocks.
"""

def build_capitup_health_excel(baseline, quotes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Health GMC Comparison"
    ws.views.sheetView[0].showGridLines = True
    
    num_columns = len(quotes) + 2
    last_col_letter = get_column_letter(num_columns)
    
    font_family = "Calibri"
    thin_side = Side(style='thin', color='000000')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    def style_range(cell_range, fill=None, font=None, alignment=None, border=cell_border):
        for row in ws[cell_range]:
            for cell in row:
                if fill: cell.fill = fill
                if font: cell.font = font
                if alignment: cell.alignment = alignment
                if border: cell.border = border

    # Header Row 1: Dark Blue Banner
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "CAPITUP INDIA PRIVATE LIMITED"
    style_range(f"A1:{last_col_letter}1",
                fill=PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
                font=Font(name=font_family, size=15, bold=True, color="FFFFFF"),
                alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28
    
    # Header Row 2: Light Green Banner
    group_name = str(baseline.get("insured_name") or "GROUP").upper()
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"HEALTH GMC COMPARISON - {group_name}"
    style_range(f"A2:{last_col_letter}2",
                fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                font=Font(name=font_family, size=11, bold=True, color="000000"),
                alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 22

    # Metadata rows
    metadata = [
        ("Corporate / Insured Group", baseline.get("insured_name", "N/A")),
        ("Policy Type", baseline.get("policy_type", "Group Medical Cover (GMC)")),
        ("Total Covered Lives (Baseline)", baseline.get("employee_count", "N/A")),
        ("Family Definition (Baseline)", baseline.get("family_definition", "N/A")),
        ("Relations Covered (Baseline)", baseline.get("relations_covered", "N/A")),
        ("Min Entry Age limit", baseline.get("min_age_limit", "N/A")),
        ("Max Entry Age limit", baseline.get("max_age_limit", "N/A")),
        ("Sum Insured Structure (Baseline)", baseline.get("sum_insured", "N/A")),
    ]
    
    current_row = 3
    for label, val in metadata:
        ws.cell(row=current_row, column=1, value=label)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=num_columns)
        ws.cell(row=current_row, column=2, value=val)
        
        ws.cell(row=current_row, column=1).font = Font(name=font_family, size=11)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        
        style_range(f"B{current_row}:{last_col_letter}{current_row}",
                    font=Font(name=font_family, size=11, bold=True),
                    alignment=Alignment(horizontal="center", vertical="center"))
        ws.cell(row=current_row, column=1).border = cell_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Insurers Heading Row
    ws.cell(row=current_row, column=1, value="Insurers").font = Font(name=font_family, size=11, bold=True)
    ws.cell(row=current_row, column=1).border = cell_border
    
    base_insurer = str(baseline.get('insurer_name') or 'EXISTING').upper()
    existing_insurer = f"EXISTING ({base_insurer})"
    ws.cell(row=current_row, column=2, value=existing_insurer).font = Font(name=font_family, size=11, bold=True)
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=current_row, column=2).border = cell_border
    
    excel_cols_written = {existing_insurer: True}
    for idx, q in enumerate(quotes):
        col_idx = idx + 3
        q_insurer = str(q.get("insurer_name") or "UNKNOWN").upper()
        
        col_name = q_insurer
        counter = 1
        while col_name in excel_cols_written:
            col_name = f"{q_insurer} ({counter})"
            counter += 1
            
        excel_cols_written[col_name] = True
        ws.cell(row=current_row, column=col_idx, value=col_name).font = Font(name=font_family, size=11, bold=True)
        ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=col_idx).border = cell_border
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Commercial values mapping
    commercials = [
        ("Gross Premium (Basic)", "gross_premium", "#,##,##0"),
        ("GST @ 18%", "gst", "#,##,##0"),
        ("Total Premium Payable", "total_premium", "#,##,##0"),
    ]
    
    for label, json_key, num_fmt in commercials:
        is_gross = (label == "Total Premium Payable")
        ws.cell(row=current_row, column=1, value=label).font = Font(name=font_family, size=11, bold=is_gross)
        ws.cell(row=current_row, column=1).border = cell_border
        
        base_val = baseline.get(json_key, 0)
        cell_base = ws.cell(row=current_row, column=2, value=base_val)
        cell_base.font = Font(name=font_family, size=11, bold=is_gross)
        cell_base.alignment = Alignment(horizontal="center", vertical="center")
        cell_base.border = cell_border
        if num_fmt and isinstance(base_val, (int, float)):
            cell_base.number_format = num_fmt
            
        for idx, q in enumerate(quotes):
            col_idx = idx + 3
            q_val = q.get(json_key, 0)
            cell_q = ws.cell(row=current_row, column=col_idx, value=q_val)
            cell_q.font = Font(name=font_family, size=11, bold=is_gross)
            cell_q.alignment = Alignment(horizontal="center", vertical="center")
            cell_q.border = cell_border
            if num_fmt and isinstance(q_val, (int, float)):
                cell_q.number_format = num_fmt
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Coverages Header row
    ws.merge_cells(f"A{current_row}:{last_col_letter}{current_row}")
    ws.cell(row=current_row, column=1, value="Coverages:").font = Font(name=font_family, size=11, bold=True)
    style_range(f"A{current_row}:{last_col_letter}{current_row}",
                font=Font(name=font_family, size=11, bold=True),
                alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # Health GMC Clauses mapping
    clauses_list = [
        ("Room Rent Limits", "Room Rent Limits"),
        ("ICU Limit", "ICU Limit"),
        ("Pre-Existing Disease waiting period", "Pre-Existing Disease waiting period"),
        ("Maternity Limit (Normal)", "Maternity Limit (Normal)"),
        ("Maternity Limit (C-Section)", "Maternity Limit (C-Section)"),
        ("Maternity Waiting Period", "Maternity Waiting Period"),
        ("Pre & Post Natal Expenses", "Pre & Post Natal Expenses"),
        ("New Born Baby Cover", "New Born Baby Cover"),
        ("Road Ambulance Limit", "Road Ambulance Limit"),
        ("Pre & Post Hospitalization Period", "Pre & Post Hospitalization Period"),
        ("AYUSH Treatment", "AYUSH Treatment"),
        ("Disease Sublimits", "Disease Sublimits"),
        ("Cataract Sublimit", "Cataract Sublimit"),
        ("Internal Congenital Disease", "Internal Congenital Disease"),
        ("External Congenital Anomaly", "External Congenital Anomaly"),
        ("FESS (Sinus Surgery)", "FESS (Sinus Surgery)"),
        ("Psychiatric & Mental Illness", "Psychiatric & Mental Illness"),
        ("Modern Treatments", "Modern Treatments"),
        ("Lasik Cover", "Lasik Cover"),
        ("Day Care Treatment", "Day Care Treatment"),
        ("Domiciliary Hospitalization", "Domiciliary Hospitalization"),
        ("Organ Donor Expenses", "Organ Donor Expenses"),
        ("Well Mother Cover", "Well Mother Cover"),
        ("Co-payment", "Co-payment"),
        ("Special Conditions", "Special Conditions")
    ]
    
    for label, json_key in clauses_list:
        ws.cell(row=current_row, column=1, value=label).font = Font(name=font_family, size=11)
        ws.cell(row=current_row, column=1).border = cell_border
        
        base_cov = baseline.get("coverages", {}).get(json_key, "No")
        cell_base = ws.cell(row=current_row, column=2, value=base_cov)
        cell_base.font = Font(name=font_family, size=11)
        cell_base.alignment = Alignment(horizontal="center", vertical="center")
        cell_base.border = cell_border
        
        for idx, q in enumerate(quotes):
            col_idx = idx + 3
            q_cov = q.get("coverages", {}).get(json_key, "No")
            cell_q = ws.cell(row=current_row, column=col_idx, value=q_cov)
            cell_q.font = Font(name=font_family, size=11)
            cell_q.alignment = Alignment(horizontal="center", vertical="center")
            cell_q.border = cell_border
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    ws.column_dimensions['A'].width = 38
    for col in range(2, num_columns + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def render_health_vertical(api_keys):
    st.subheader("🏥 Health Insurance (GMC) Analyzer")
    st.markdown("Upload Health GMC Quotes. The layout parser will extract benefits, waiting periods, and room rent structures side-by-side.")
    
    col_a, col_b = st.columns(2)
    with col_a:
        prev_health_file = st.file_uploader("1. Previous Year GMC Policy / RFQ Slip", type=["pdf"])
    with col_b:
        new_health_files = st.file_uploader("2. Current GMC Quotations", type=["pdf"], accept_multiple_files=True)
        
    if prev_health_file and new_health_files:
        if not api_keys:
            st.error("Please configure API keys in the sidebar or `.streamlit/secrets.toml`.")
            return
            
        if st.button("Generate Health GMC QCR Matrix", type="primary", width="stretch"):
            baseline_health = None
            health_records = []
            
            try:
                key_manager = APIKeyManager(api_keys)
                
                with st.spinner("Extracting GMC baseline details..."):
                    prev_bytes = prev_health_file.read()
                    baseline_health = process_pdf_document(prev_bytes, prev_health_file.name, HEALTH_SYSTEM_PROMPT, key_manager)

                for q_file in new_health_files:
                    with st.spinner(f"Extracting GMC details from {q_file.name}..."):
                        q_bytes = q_file.read()
                        record = process_pdf_document(q_bytes, q_file.name, HEALTH_SYSTEM_PROMPT, key_manager)
                        health_records.append(record)
                
                if baseline_health and health_records:
                    st.session_state.health_baseline = baseline_health
                    st.session_state.health_records = health_records
                    st.success("Health GMC analysis complete.")
            
            except Exception as e:
                st.error(f"Processing halted: {e}")

    # Render Health GMC Preview Table
    if "health_baseline" in st.session_state and "health_records" in st.session_state:
        hb = st.session_state.health_baseline
        h_records = st.session_state.health_records
        
        excel_bytes = build_capitup_health_excel(hb, h_records)
        
        columns_map = {}
        
        # 1. Populate baseline columns with deduplication
        base_insurer = str(hb.get('insurer_name') or 'EXISTING').upper()
        base_key = f"EXISTING ({base_insurer})"
        columns_map[base_key] = [
            hb.get("insured_name"),
            hb.get("policy_type", "Group Medical Cover (GMC)"),
            hb.get("employee_count"),
            hb.get("family_definition"),
            hb.get("relations_covered"),
            f"Min: {hb.get('min_age_limit', 'N/A')} | Max: {hb.get('max_age_limit', 'N/A')}",
            hb.get("sum_insured"),
            hb.get("gross_premium"),
            hb.get("gst"),
            hb.get("total_premium"),
            hb.get("coverages", {}).get("Room Rent Limits", "No Capping"),
            hb.get("coverages", {}).get("ICU Limit", "No Capping"),
            hb.get("coverages", {}).get("Pre-Existing Disease waiting period", "No"),
            hb.get("coverages", {}).get("Maternity Limit (Normal)", "No"),
            hb.get("coverages", {}).get("Maternity Limit (C-Section)", "No"),
            hb.get("coverages", {}).get("Maternity Waiting Period", "No"),
            hb.get("coverages", {}).get("Pre & Post Natal Expenses", "No"),
            hb.get("coverages", {}).get("New Born Baby Cover", "No"),
            hb.get("coverages", {}).get("Road Ambulance Limit", "No"),
            hb.get("coverages", {}).get("Pre & Post Hospitalization Period", "No"),
            hb.get("coverages", {}).get("AYUSH Treatment", "No"),
            hb.get("coverages", {}).get("Disease Sublimits", "No"),
            hb.get("coverages", {}).get("Cataract Sublimit", "No"),
            hb.get("coverages", {}).get("Internal Congenital Disease", "No"),
            hb.get("coverages", {}).get("External Congenital Anomaly", "No"),
            hb.get("coverages", {}).get("FESS (Sinus Surgery)", "No"),
            hb.get("coverages", {}).get("Psychiatric & Mental Illness", "No"),
            hb.get("coverages", {}).get("Modern Treatments", "No"),
            hb.get("coverages", {}).get("Lasik Cover", "No"),
            hb.get("coverages", {}).get("Day Care Treatment", "No"),
            hb.get("coverages", {}).get("Domiciliary Hospitalization", "No"),
            hb.get("coverages", {}).get("Organ Donor Expenses", "No"),
            hb.get("coverages", {}).get("Well Mother Cover", "No"),
            hb.get("coverages", {}).get("Co-payment", "No"),
            hb.get("coverages", {}).get("Special Conditions", "No")
        ]
        
        # 2. Populate Quotes dynamically
        for idx, rec in enumerate(h_records):
            insurer = str(rec.get("insurer_name") or f"Insurer {idx+1}").upper()
            
            col_name = insurer
            counter = 1
            while col_name in columns_map:
                col_name = f"{insurer} ({counter})"
                counter += 1
                
            columns_map[col_name] = [
                rec.get("insured_name"),
                rec.get("policy_type", "Group Medical Cover (GMC)"),
                rec.get("employee_count"),
                rec.get("family_definition"),
                rec.get("relations_covered"),
                f"Min: {rec.get('min_age_limit', 'N/A')} | Max: {rec.get('max_age_limit', 'N/A')}",
                rec.get("sum_insured"),
                rec.get("gross_premium"),
                rec.get("gst"),
                rec.get("total_premium"),
                rec.get("coverages", {}).get("Room Rent Limits", "No Capping"),
                rec.get("coverages", {}).get("ICU Limit", "No Capping"),
                rec.get("coverages", {}).get("Pre-Existing Disease waiting period", "No"),
                rec.get("coverages", {}).get("Maternity Limit (Normal)", "No"),
                rec.get("coverages", {}).get("Maternity Limit (C-Section)", "No"),
                rec.get("coverages", {}).get("Maternity Waiting Period", "No"),
                rec.get("coverages", {}).get("Pre & Post Natal Expenses", "No"),
                rec.get("coverages", {}).get("New Born Baby Cover", "No"),
                rec.get("coverages", {}).get("Road Ambulance Limit", "No"),
                rec.get("coverages", {}).get("Pre & Post Hospitalization Period", "No"),
                rec.get("coverages", {}).get("AYUSH Treatment", "No"),
                rec.get("coverages", {}).get("Disease Sublimits", "No"),
                rec.get("coverages", {}).get("Cataract Sublimit", "No"),
                rec.get("coverages", {}).get("Internal Congenital Disease", "No"),
                rec.get("coverages", {}).get("External Congenital Anomaly", "No"),
                rec.get("coverages", {}).get("FESS (Sinus Surgery)", "No"),
                rec.get("coverages", {}).get("Psychiatric & Mental Illness", "No"),
                rec.get("coverages", {}).get("Modern Treatments", "No"),
                rec.get("coverages", {}).get("Lasik Cover", "No"),
                rec.get("coverages", {}).get("Day Care Treatment", "No"),
                rec.get("coverages", {}).get("Domiciliary Hospitalization", "No"),
                rec.get("coverages", {}).get("Organ Donor Expenses", "No"),
                rec.get("coverages", {}).get("Well Mother Cover", "No"),
                rec.get("coverages", {}).get("Co-payment", "No"),
                rec.get("coverages", {}).get("Special Conditions", "No")
            ]
            
        index_labels = [
            "Corporate / Insured Group", "Policy Type", "Total Covered Lives", "Family Definition", "Relations Covered", "Entry Age Limits", "Sum Insured Structure",
            "Gross Premium (Basic)", "GST @ 18%", "Total Premium Payable",
            "Room Rent Limits", "ICU Limits", "Pre-Existing Disease Waiver", "Maternity Limits (Normal)", "Maternity Limits (C-Section)", "Maternity Waiting Period",
            "Pre & Post Natal Expenses", "New Born Baby Cover", "Road Ambulance Limit", "Pre & Post Hospitalization Period", "AYUSH Treatment",
            "Disease Sublimits", "Cataract Sublimit", "Internal Congenital Coverage", "External Congenital Coverage", "FESS (Sinus Surgery)",
            "Psychiatric & Mental Illness", "Modern Treatment Cover", "Lasik Cover", "Day Care Treatment", "Domiciliary Hospitalization", "Organ Donor Expenses",
            "Well Mother Cover", "Co-payment Clauses", "Special Conditions"
        ]
        
        df_health_preview = pd.DataFrame(columns_map, index=index_labels).astype(str)
        
        st.write("---")
        st.subheader("📋 Health GMC QCR Matrix Preview")
        st.dataframe(df_health_preview, width="stretch")
        
        st.download_button(
            label="📥 Download Health GMC QCR Matrix (Excel)",
            data=excel_bytes,
            file_name="GMC_Comparison_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch"
        )
        
        # Display math audit panel at the very bottom
        render_underwriter_audit_panel(hb, h_records)
