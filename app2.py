import streamlit as st
import tempfile
import os
import json
import io
import time
import pandas as pd
from datetime import datetime

# Import the modern Google GenAI SDK
from google import genai
from google.genai import types, errors

# Import styling components for Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set up page configurations
st.set_page_config(page_title="CapitUp QCR Engine", page_icon="🛡️", layout="wide")

# ==========================================
# NATIVE STREAMLIT SECRETS CONFIGURATION
# ==========================================
default_keys = []
try:
    if "GEMINI_API_KEYS" in st.secrets:
        secret_data = st.secrets["GEMINI_API_KEYS"]
        if isinstance(secret_data, list):
            default_keys = [str(k).strip() for k in secret_data if str(k).strip()]
        elif isinstance(secret_data, str):
            default_keys = [k.strip() for k in secret_data.split(',') if k.strip()]
except Exception:
    pass

st.sidebar.title("CapitUp Systems")

default_keys_str = ",".join(default_keys) if default_keys else ""

global_api_keys = st.sidebar.text_input(
    "API Keys Configuration (Comma-separated)",
    value=default_keys_str,
    type="password",
    help="Loaded automatically from .streamlit/secrets.toml if configured."
)

if default_keys:
    st.sidebar.success(f"✅ Auto-loaded {len(default_keys)} keys from secrets!")

# ==========================================
# ENTERPRISE API KEY MANAGER
# ==========================================
class APIKeyManager:
    def __init__(self, keys_input):
        if isinstance(keys_input, list):
            self.keys = [str(k).strip() for k in keys_input if str(k).strip()]
        elif isinstance(keys_input, str):
            self.keys = [k.strip() for k in keys_input.split(',') if k.strip()]
        else:
            self.keys = []
            
        self.current_index = 0
        self.dead_keys = set() 

    def get_next_key(self):
        if not self.keys:
            raise ValueError("No API keys provided.")
        if len(self.dead_keys) == len(self.keys):
            raise Exception("All provided API keys are permanently invalid (401/403).")
        
        start_index = self.current_index
        while True:
            key = self.keys[self.current_index]
            self.current_index = (self.current_index + 1) % len(self.keys)
            
            if key not in self.dead_keys:
                return key
                
            if self.current_index == start_index:
                raise Exception("All keys are permanently dead.")

    def mark_key_dead(self, key, reason):
        masked_key = f"{key[:6]}...{key[-4:]}" if len(key) > 10 else "UNKNOWN_KEY"
        st.sidebar.error(f"💀 Key ({masked_key}) permanently invalid: {reason}.")
        self.dead_keys.add(key)


# ==========================================
# CORE AI PROCESSING & MODEL WATERFALL
# ==========================================
def process_pdf_document(file_bytes, file_name, prompt, key_manager, retries=6):
    # Standardized with active production model names to avoid 404 errors
    model_waterfall = ['gemini-3.6-flash','gemini-3.5-flash', 'gemini-3.1-pro-preview','gemini-3-pro-preview','gemini-2.5-flash', 'gemini-2.5-flash-lite']
    
    json_config = types.GenerateContentConfig(
        response_mime_type="application/json",
        temperature=0.0
    )

    for attempt in range(retries):
        current_model = model_waterfall[attempt % len(model_waterfall)]
        
        try:
            api_key = key_manager.get_next_key()
            client = genai.Client(api_key=api_key)
            
            response = client.models.generate_content(
                model=current_model,
                contents=[
                    types.Part.from_bytes(data=file_bytes, mime_type='application/pdf'),
                    prompt
                ], 
                config=json_config
            )
            
            data = json.loads(response.text.strip())
            return data
            
        except errors.APIError as e:
            if e.code == 429:
                masked = f"{api_key[:6]}...{api_key[-4:]}" if len(api_key) > 10 else "KEY"
                st.warning(f"⏳ Rate Limit on {current_model} (Key {masked}). Cooling down for 15s... (Attempt {attempt + 1}/{retries})")
                time.sleep(15) 
            elif e.code in [400, 401, 403]:
                key_manager.mark_key_dead(api_key, f"Auth Error ({e.code})")
            elif e.code >= 500:
                st.warning(f"🔄 Google Server Error ({e.code}) on {current_model}. Swapping to fallback model...")
                time.sleep(2)
            else:
                st.error(f"API Error on {file_name}: {e.message}")
                time.sleep(5)
                
        except json.JSONDecodeError:
            st.warning(f"🔄 {current_model} failed to format JSON on {file_name}. Swapping to fallback model...")
            
        except Exception as e:
            if "permanently dead" in str(e):
                raise e 
            st.error(f"Unexpected error on {file_name}: {str(e)}")
            time.sleep(2)
            
    raise Exception(f"Failed to process {file_name} after {retries} attempts across all models.")


# ==========================================
# EXCEL GENERATOR (Capitup Branded Motor QCR)
# ==========================================
def build_capitup_motor_excel(baseline, quotes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.views.sheetView[0].showGridLines = True
    
    num_columns = len(quotes) + 2
    last_col_letter = get_column_letter(num_columns)
    
    font_family = "Calibri"
    thin_side = Side(style='thin', color='000000')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    def style_range(cell_range, fill=None, font=None, alignment=None, border=cell_border):
        for row in ws[cell_range]:
            for cell in row:
                if fill: cell.fill = fill
                if font: cell.font = font
                if alignment: cell.alignment = alignment
                if border: cell.border = border

    # Header Row 1: Dark Blue Banner
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "CAPITUP INDIA PRIVATE LIMITED"
    style_range(f"A1:{last_col_letter}1",
                fill=PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
                font=Font(name=font_family, size=15, bold=True, color="FFFFFF"),
                alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28
    
    # Header Row 2: Light Green Banner
    make = str(baseline.get("make") or "VEHICLE").upper()
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"COMPARISON - {make}"
    style_range(f"A2:{last_col_letter}2",
                fill=PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                font=Font(name=font_family, size=11, bold=True, color="000000"),
                alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 22

    metadata = [
        ("Name of the Insured", baseline.get("insured_name", "N/A")),
        ("Address", baseline.get("address", "N/A")),
        ("Renewal date", baseline.get("renewal_date", "N/A")),
        ("Registration Number", baseline.get("registration_number", "N/A")),
        ("Make", baseline.get("make", "N/A")),
        ("Model", baseline.get("model", "N/A")),
        ("Seating Capacity", baseline.get("seating_capacity", "N/A")),
        ("Cubic Capacity", baseline.get("cubic_capacity", "N/A")),
        ("Year of Manufacture", baseline.get("year_of_manufacture", "N/A")),
    ]
    
    current_row = 3
    for label, val in metadata:
        ws.cell(row=current_row, column=1, value=label)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=num_columns)
        ws.cell(row=current_row, column=2, value=val)
        
        ws.cell(row=current_row, column=1).font = Font(name=font_family, size=11)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        
        style_range(f"B{current_row}:{last_col_letter}{current_row}",
                    font=Font(name=font_family, size=11, bold=True),
                    alignment=Alignment(horizontal="center", vertical="center"))
        ws.cell(row=current_row, column=1).border = cell_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Insurers Heading Row
    ws.cell(row=current_row, column=1, value="Insurers").font = Font(name=font_family, size=11, bold=True)
    ws.cell(row=current_row, column=1).border = cell_border
    
    base_insurer = str(baseline.get('insurer_name') or 'TATA')
    existing_insurer = f"EXISTING ({base_insurer})"
    ws.cell(row=current_row, column=2, value=existing_insurer).font = Font(name=font_family, size=11, bold=True)
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=current_row, column=2).border = cell_border
    
    # Store dynamic quote names to ensure unique Excel column headers
    excel_cols_written = {existing_insurer: True}
    for idx, q in enumerate(quotes):
        col_idx = idx + 3
        q_insurer = str(q.get("insurer_name") or "UNKNOWN").upper()
        
        col_name = q_insurer
        counter = 1
        while col_name in excel_cols_written:
            col_name = f"{q_insurer} ({counter})"
            counter += 1
            
        excel_cols_written[col_name] = True
        ws.cell(row=current_row, column=col_idx, value=col_name).font = Font(name=font_family, size=11, bold=True)
        ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=col_idx).border = cell_border
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    commercials = [
        ("IDV", "idv", "#,##,##0"),
        ("NCB in %", "ncb_percent", None),
        ("TP", "tp_premium", "#,##,##0"),
        ("OD Premium", "od_premium", "#,##,##0"),
        ("Gst @ 18%", "gst", "#,##,##0"),
        ("Gross Premium", "gross_premium", "#,##,##0"),
    ]
    
    for label, json_key, num_fmt in commercials:
        is_gross = (label == "Gross Premium")
        ws.cell(row=current_row, column=1, value=label).font = Font(name=font_family, size=11, bold=is_gross)
        ws.cell(row=current_row, column=1).border = cell_border
        
        base_val = baseline.get(json_key, 0)
        cell_base = ws.cell(row=current_row, column=2, value=base_val)
        cell_base.font = Font(name=font_family, size=11, bold=is_gross)
        cell_base.alignment = Alignment(horizontal="center", vertical="center")
        cell_base.border = cell_border
        if num_fmt and isinstance(base_val, (int, float)):
            cell_base.number_format = num_fmt
            
        for idx, q in enumerate(quotes):
            col_idx = idx + 3
            q_val = q.get(json_key, 0)
            cell_q = ws.cell(row=current_row, column=col_idx, value=q_val)
            cell_q.font = Font(name=font_family, size=11, bold=is_gross)
            cell_q.alignment = Alignment(horizontal="center", vertical="center")
            cell_q.border = cell_border
            if num_fmt and isinstance(q_val, (int, float)):
                cell_q.number_format = num_fmt
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    ws.merge_cells(f"A{current_row}:{last_col_letter}{current_row}")
    ws.cell(row=current_row, column=1, value="Coverages:").font = Font(name=font_family, size=11, bold=True)
    style_range(f"A{current_row}:{last_col_letter}{current_row}",
                font=Font(name=font_family, size=11, bold=True),
                alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    all_coverage_keys = set()
    all_coverage_keys.update(baseline.get("coverages", {}).keys())
    for q in quotes:
        all_coverage_keys.update(q.get("coverages", {}).keys())
        
    standard_order = [
        "Zero Depreciation", "Engine Protection", "Return to Invoice", "Tyre Protection",
        "Consumables", "Roadside Assistance", "Key Replacement", "Passenger Assistance",
        "Loss of Personal Belongings", "Basic TP", "PA Cover", "Legal Liability to Paid Driver"
    ]
    
    ordered_keys = [k for k in standard_order if k in all_coverage_keys]
    dynamic_keys = sorted([k for k in all_coverage_keys if k not in standard_order])
    final_coverage_rows = ordered_keys + dynamic_keys
    
    for key in final_coverage_rows:
        ws.cell(row=current_row, column=1, value=key).font = Font(name=font_family, size=11)
        ws.cell(row=current_row, column=1).border = cell_border
        
        base_cov = baseline.get("coverages", {}).get(key, "No")
        cell_base = ws.cell(row=current_row, column=2, value=base_cov)
        cell_base.font = Font(name=font_family, size=11)
        cell_base.alignment = Alignment(horizontal="center", vertical="center")
        cell_base.border = cell_border
        
        for idx, q in enumerate(quotes):
            col_idx = idx + 3
            q_cov = q.get("coverages", {}).get(key, "No")
            cell_q = ws.cell(row=current_row, column=col_idx, value=q_cov)
            cell_q.font = Font(name=font_family, size=11)
            cell_q.alignment = Alignment(horizontal="center", vertical="center")
            cell_q.border = cell_border
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    ws.column_dimensions['A'].width = 30
    for col in range(2, num_columns + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), final_coverage_rows


# ==========================================
# UNDERWRITER MATHEMATICAL AUDIT ENGINE
# ==========================================
class CommercialReconciliationEngine:
    def __init__(self, data_record):
        self.data = data_record
        self.insurer = str(data_record.get("insurer_name") or "UNKNOWN").upper()
        self.gross_premium = float(str(data_record.get("gross_premium") or 0).replace(",", ""))
        self.od_premium = float(str(data_record.get("od_premium") or 0).replace(",", ""))
        self.tp_premium = float(str(data_record.get("tp_premium") or 0).replace(",", ""))
        self.gst = float(str(data_record.get("gst") or 0).replace(",", ""))
        
    def run_reconciliation_audit(self):
        audit_results = {
            "status": "PASSED",
            "logs": [],
            "reconciliation_delta": 0.0
        }
        
        calculated_gross = self.od_premium + self.tp_premium + self.gst
        reconciliation_delta = abs(self.gross_premium - calculated_gross)
        
        audit_results["logs"].append(
            f"ℹ️ **Document Values**: Net OD Premium = ₹{self.od_premium:,.2f} | "
            f"TP Premium = ₹{self.tp_premium:,.2f} | GST = ₹{self.gst:,.2f}"
        )
        audit_results["logs"].append(
            f"ℹ️ **Total Sum**: Calculated sum is ₹{calculated_gross:,.2f} vs. "
            f"Document printed Gross: ₹{self.gross_premium:,.2f}"
        )
        
        if reconciliation_delta <= 2.0:
            audit_results["logs"].append(
                f"✅ **Premium Reconciliation Passed**: Formula matches Gross exactly. "
                f"Variance: ₹{reconciliation_delta:.2f}"
            )
        else:
            audit_results["status"] = "WARNING"
            audit_results["reconciliation_delta"] = reconciliation_delta
            audit_results["logs"].append(
                f"⚠️ **Premium Deviation Notice**: Sum of extracted components equals calculated ₹{calculated_gross:,.2f}, "
                f"while the document lists Gross as ₹{self.gross_premium:,.2f}. "
                f"Deviation: **₹{reconciliation_delta:,.2f}** (This can occur due to rounding or bundled parameters)."
            )
            
        return audit_results

def render_underwriter_audit_panel(baseline_record, quotes_records):
    st.markdown("---")
    st.subheader("🛡️ Underwriter Audit & Mathematical Reconciliation Logs")
    st.info("Reconciliation checks run automatically in the background to analyze the extracted values.")
    
    # Baseline audit
    base_engine = CommercialReconciliationEngine(baseline_record)
    base_audit = base_engine.run_reconciliation_audit()
    
    with st.expander(f"🔍 Commercial Audit Log: EXISTING ({base_engine.insurer})", expanded=False):
        for log in base_audit["logs"]:
            st.markdown(log)
            
    # Quotes audit
    for q in quotes_records:
        q_engine = CommercialReconciliationEngine(q)
        q_audit = q_engine.run_reconciliation_audit()
        
        with st.expander(f"🔍 Commercial Audit Log: {q_engine.insurer}", expanded=False):
            for log in q_audit["logs"]:
                st.markdown(log)


# ==========================================
# MASTER LAYOUT DEFINITION & Vertical ROUTER
# ==========================================
def main():
    domain = st.sidebar.selectbox(
        "Select Insurance Domain", 
        ["🚗 Motor QCR Generator", "🏥 Health Insurance (GMC)"]
    )
    
    st.sidebar.write("---")
    
    if domain == "🚗 Motor QCR Generator":
        render_motor_vertical(global_api_keys)
    elif domain == "🏥 Health Insurance (GMC)":
        render_health_vertical(global_api_keys)


# ==========================================
# MOTOR INSURANCE MODULE
# ==========================================
MOTOR_SYSTEM_PROMPT = """
You are an expert, elite-level motor insurance underwriter and auditor. Your task is to extract variables from the provided document (policy schedule, quote, or draft) with absolute analytical precision.

*** STRICT UNDERWRITING RULES FOR DYNAMIC ACCURACY (READ CAREFULLY) ***

1. THE ZERO-PREMIUM RULE (CRITICAL): Policy schedules and quotes often print a static template listing every available add-on cover. 
   - If an add-on (e.g. "Tyre Protection", "Key Replacement", "Consumables", "Repair of Glass, Rubber & Plastic Parts (TA08)") is listed with a premium of "0", "0.00", "-", "Nil", "N/A", or is blank/dash, you MUST classify it as "No" (NOT covered).
   - This applies even if the add-on name is clearly printed on the page. No premium paid = No coverage.

2. EXCEPTION TO THE ZERO-PREMIUM RULE: You may only classify a zero-premium add-on as "Yes" if there is explicit visual/textual proof of inclusion adjacent to the add-on name inside the active premium table, such as:
   - A clear checkmark (✓), or explicitly written words like "Included", "Inbuilt", "FOC" (Free of Cost), "Complimentary".
   - The premium value is clearly integrated into another package bundle explicitly documented in the schedule (e.g. ICICI's "Prestige" bundle).

3. NO EXTRACTING FROM MARKETING TEXT / USPs: 
   - DO NOT extract coverage status from marketing boxes, promotional footers, or sales pitches.
   - For example, if the page has a "Why Choose Us / Our USP" sidebar or footer (like Liberty's "24x7 liberty complete assistance" USP box) that checkmarks features like "24x7 roadside assistance", but Roadside Assistance has a premium of 0.00 and is not itemized as active in the premium calculation table, you MUST classify it as "No". 
   - However, for Liberty specifically, Roadside Assistance is considered active ("Yes") ONLY if the "24x7 liberty complete assistance" box is explicitly checked at the bottom of their page.
   - For ICICI Lombard, "Roadside Assistance" is classified as "No" because "IL Smart Assist" is their carrier-specific add-on. Extract "IL Smart Assist" as a separate row ("Yes") and keep "Roadside Assistance" strictly as "No" to match the broker's preference.

4. CPA (Compulsory Personal Accident) WAIVER: If the CPA premium is 0, verify if an owner-driver waiver has been selected. If opted out, "PA Cover" is strictly "No".

5. ANCHORING AUDIT TRAIL: In your JSON response, you must populate the "audit_trail" object. Explain precisely where on the document (page, table, line) you verified the IDV, the gross premium, and why each coverage is marked "Yes" or "No" based on the premium details. This forces your processing to be 100% correct.

Extract and format the output as a clean JSON object matching this schema:
{
  "insurer_name": "Name of insurer, e.g., TATA AIG, ICICI LOMBARD, LIBERTY",
  "insured_name": "Name of Insured / Policyholder",
  "address": "State or Location of Address",
  "renewal_date": "Policy expiration / renewal date in DD-MM-YYYY format",
  "registration_number": "Vehicle Registration Number",
  "make": "Vehicle Make (e.g. TOYOTA)",
  "model": "Vehicle Model & Variant (e.g. INNOVA HYCROSS)",
  "seating_capacity": 5,
  "cubic_capacity": 1987,
  "year_of_manufacture": 2024,
  "idv": 2400000,
  "ncb_percent": "20%",
  "tp_premium": 0.0,
  "od_premium": 29362.0,
  "gst": 5285.0,
  "gross_premium": 34647.0,
  "coverages": {
    "Zero Depreciation": "Yes or No",
    "Engine Protection": "Yes or No",
    "Return to Invoice": "Yes or No",
    "Tyre Protection": "Yes or No",
    "Consumables": "Yes or No",
    "Roadside Assistance": "Yes or No",
    "Key Replacement": "Yes or No",
    "Passenger Assistance": "Yes or No",
    "Loss of Personal Belongings": "Yes or No",
    "Basic TP": "Yes or No",
    "PA Cover": "Yes or No",
    "Legal Liability to Paid Driver": "Yes or No"
  },
  "audit_trail": {
    "idv_location": "Explicit location / text proving IDV value",
    "premium_cross_check": "Formula verifying: Net OD + TP + GST = Gross Premium",
    "zero_premium_validation_notes": "Validation audit logs of why zero premium add-ons were flagged as Yes/No"
  }
}

Use standard normalizations for coverage keys:
- Depreciation Reimbursement / Nil Depreciation / Zero Dep -> "Zero Depreciation"
- Engine Protect / Engine Secure / Engine Safe -> "Engine Protection"
- Loss of Personal Belongings -> "Loss of Personal Belongings"
- Passenger Assistance / Emergency Assistance -> "Passenger Assistance"
- Key Replacement / Key Loss -> "Key Replacement"
- Consumable Expenses / Consumables -> "Consumables"
- Road side Assistance / RSA / IL Smart Assist -> "Roadside Assistance"
- Gap Value / Return to Invoice / Gap Value 1 -> "Return to Invoice"
- Tyre Secure / Tyre Protection -> "Tyre Protection"

If any other unique add-on cover (e.g. "Liberty Assure", "Smart Saver", "Emergency Medical Expenses") is explicitly active, add it to the coverages dictionary under its clean name as "Yes".

Ensure you return ONLY valid JSON. Do not write any markdown blocks.
"""

def render_motor_vertical(api_keys):
    st.subheader("🚗 Motor Insurance QCR Generation Module")
    st.markdown("Process previous year schedules and new comparative quotes directly using Layout OCR.")
    
    col_a, col_b = st.columns(2)
    with col_a:
        prev_file = st.file_uploader("1. Previous Year Policy Copy", type=["pdf"])
    with col_b:
        new_files = st.file_uploader("2. Current Quotations", type=["pdf"], accept_multiple_files=True)
        
    if prev_file and new_files:
        if not api_keys:
            st.error("Please configure API keys in the sidebar or `.streamlit/secrets.toml`.")
            return
            
        if st.button("Generate Motor QCR Report", type="primary", width="stretch"):
            baseline = None
            quotes = []
            
            try:
                key_manager = APIKeyManager(api_keys)
                
                with st.spinner("Analyzing Previous Policy details..."):
                    prev_bytes = prev_file.read()
                    baseline = process_pdf_document(prev_bytes, prev_file.name, MOTOR_SYSTEM_PROMPT, key_manager)

                for q_file in new_files:
                    with st.spinner(f"Analyzing quote: {q_file.name}..."):
                        q_bytes = q_file.read()
                        quote_data = process_pdf_document(q_bytes, q_file.name, MOTOR_SYSTEM_PROMPT, key_manager)
                        quotes.append(quote_data)
                            
                if baseline and quotes:
                    st.session_state.motor_baseline = baseline
                    st.session_state.motor_quotes = quotes
                    st.success("Comparison extraction completed successfully.")
                    
            except Exception as e:
                st.error(f"Processing halted: {e}")

    # Render Preview, Excel Generation, and Reconciliation logs
    if "motor_baseline" in st.session_state and "motor_quotes" in st.session_state:
        b = st.session_state.motor_baseline
        qs = st.session_state.motor_quotes
        
        excel_bytes, dynamic_covers = build_capitup_motor_excel(b, qs)
        
        base_insurer = str(b.get('insurer_name') or 'TATA').upper()
        
        # Build UI Dataframe Preview Map with DEDUPLICATION
        preview_cols = {}
        
        # 1. Populate baseline
        base_key = f"EXISTING ({base_insurer})"
        preview_cols[base_key] = [
            b.get("insured_name"), b.get("address"), b.get("renewal_date"), b.get("registration_number"),
            b.get("make"), b.get("model"), b.get("seating_capacity"), b.get("cubic_capacity"), b.get("year_of_manufacture"),
            "---", b.get("idv"), b.get("ncb_percent"), b.get("tp_premium"), b.get("od_premium"), b.get("gst"), b.get("gross_premium"),
            "---"
        ] + [b.get("coverages", {}).get(key, "No") for key in dynamic_covers]
        
        # 2. Populate Quotes dynamically with dynamic key deduplication to prevent PyArrow crashes on multiple matching insurers
        for q in qs:
            q_insurer = str(q.get("insurer_name") or "UNKNOWN").upper()
            
            col_name = q_insurer
            counter = 1
            while col_name in preview_cols:
                col_name = f"{q_insurer} ({counter})"
                counter += 1
                
            preview_cols[col_name] = [
                b.get("insured_name"), b.get("address"), b.get("renewal_date"), b.get("registration_number"),
                b.get("make"), b.get("model"), b.get("seating_capacity"), b.get("cubic_capacity"), b.get("year_of_manufacture"),
                "---", q.get("idv"), q.get("ncb_percent"), q.get("tp_premium"), q.get("od_premium"), q.get("gst"), q.get("gross_premium"),
                "---"
            ] + [q.get("coverages", {}).get(key, "No") for key in dynamic_covers]
            
        row_labels = [
            "Name of the Insured", "Address", "Renewal date", "Registration Number", 
            "Make", "Model", "Seating Capacity", "Cubic Capacity", "Year of Manufacture",
            "--- COMMERCIALS ---", "IDV", "NCB in %", "TP", "OD Premium", "Gst @ 18%", "Gross Premium",
            "--- COVERAGES ---"
        ] + dynamic_covers
        
        df_preview = pd.DataFrame(preview_cols, index=row_labels).astype(str)
        
        st.markdown("---")
        st.subheader("📋 Capitup Comparison Preview")
        st.dataframe(df_preview, width="stretch")
        
        st.download_button(
            label="📥 Download Capitup Comparative Sheet (Excel)",
            data=excel_bytes,
            file_name=f"Capitup_Motor_QCR_{b.get('registration_number', 'Report')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch"
        )
        
        # Display the informational logs at the very bottom
        render_underwriter_audit_panel(b, qs)


# ==========================================
# HEALTH INSURANCE MODULE (GMC QCR)
# ==========================================
HEALTH_SYSTEM_PROMPT = """
You are an expert corporate health insurance underwriter and auditor. Analyze the GMC (Group Medical Cover) quotation precisely.

*** CRITICAL RULES FOR GMC COVERAGE LIMITS ***
1. THE CO-PAYMENT RULE: If co-payment is "Nil", "No", "Waived", or "0%", classify as "No Co-pay". If a percentage is listed (e.g., "10% on claims"), write "10% Co-pay".
2. MATERNITY CLAUSE: Look for maternity benefits. If maternity is "Not Covered" or has a limit of "0", classify as "No" or "Not Covered".
3. ROOM RENT LIMITS: Check if room rent limits are capped (e.g., "1% of Sum Insured") or uncapped ("Single Private AC Room" or "No Limit").
4. ANCHORING AUDIT TRAIL: In your JSON response, you must populate the "audit_trail" object. Explain precisely where on the document (page, table, line) you verified the premium breakdown and coverage limits.

Extract and format the output as a clean JSON matching the following schema.

Extract and format the output as a clean JSON object:
{
  "insurer_name": "Name of insurer, e.g., Star Health, Care Health, Niva Bupa",
  "insured_name": "Name of Corporate Insured / Group",
  "policy_type": "Group Medical Cover (GMC)",
  "total_lives": 250,
  "sum_insured": "Family Floater of 3,00,000",
  "gross_premium": 150000.0,
  "gst": 27000.0,
  "total_premium": 177000.0,
  "coverages": {
    "Room Rent Limits": "1% of Sum Insured or Single Private AC Room",
    "ICU Limit": "2% of Sum Insured or No Limit",
    "Pre-Existing Disease waiting period": "Covered from Day 1 / Waiver",
    "Maternity Limit": "50,000 for Normal, 75,000 for Caesarean",
    "Corporate Buffer": "No / Yes (e.g. 5,00,000)",
    "Co-payment": "No Co-pay / 10% on claims"
  },
  "audit_trail": {
    "premium_location": "Page X, Table Y proving premium details",
    "maternity_verification": "Page X, Table Y proving maternity limits"
  }
}
Respond ONLY with raw JSON. No markdown wrappers.
"""

def render_health_vertical(api_keys):
    st.subheader("🏥 Health Insurance (GMC) Analyzer")
    st.markdown("Upload Health GMC Quotes. The layout parser will extract benefits, waiting periods, and room rent structures side-by-side.")
    
    uploaded_quotes = st.file_uploader("Upload GMC Quotes (PDF)", type=["pdf"], accept_multiple_files=True)
    
    if uploaded_quotes:
        if not api_keys:
            st.error("Please configure API keys in the sidebar or `.streamlit/secrets.toml`.")
            return
            
        if st.button("Generate Health GMC QCR Matrix", type="primary", width="stretch"):
            health_records = []
            
            try:
                key_manager = APIKeyManager(api_keys)
                
                for q_file in uploaded_quotes:
                    with st.spinner(f"Extracting GMC details from {q_file.name}..."):
                        q_bytes = q_file.read()
                        record = process_pdf_document(q_bytes, q_file.name, HEALTH_SYSTEM_PROMPT, key_manager)
                        health_records.append(record)
                
                if health_records:
                    st.session_state.health_records = health_records
                    st.success("Health GMC analysis complete.")
            
            except Exception as e:
                st.error(f"Processing halted: {e}")

    if "health_records" in st.session_state and st.session_state.health_records:
        h_records = st.session_state.health_records
        
        columns_map = {}
        for idx, rec in enumerate(h_records):
            insurer = str(rec.get("insurer_name") or f"Insurer {idx+1}").upper()
            columns_map[insurer] = [
                rec.get("insured_name"),
                rec.get("policy_type"),
                rec.get("total_lives"),
                rec.get("sum_insured"),
                rec.get("gross_premium"),
                rec.get("gst"),
                rec.get("total_premium"),
                rec.get("coverages", {}).get("Room Rent Limits", "No"),
                rec.get("coverages", {}).get("ICU Limit", "No"),
                rec.get("coverages", {}).get("Pre-Existing Disease waiting period", "No"),
                rec.get("coverages", {}).get("Maternity Limit", "No"),
                rec.get("coverages", {}).get("Corporate Buffer", "No"),
                rec.get("coverages", {}).get("Co-payment", "No")
            ]
            
        index_labels = [
            "Corporate / Insured Group", "Policy Type", "Total Covered Lives", "Sum Insured Structure",
            "Gross Premium (Basic)", "GST @ 18%", "Total Premium Payable",
            "Room Rent Limits", "ICU Limits", "Pre-Existing Disease Waiver", "Maternity Limits",
            "Corporate Buffer", "Co-payment Clauses"
        ]
        
        df_health_preview = pd.DataFrame(columns_map, index=index_labels).astype(str)
        
        st.write("---")
        st.subheader("📋 Health GMC QCR Matrix Preview")
        st.dataframe(df_health_preview, width="stretch")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Health GMC Comparison"
        
        ws.cell(row=1, column=1, value="Health GMC Comparison").font = Font(name="Calibri", size=14, bold=True)
        for r_idx, label in enumerate(index_labels, start=3):
            ws.cell(row=r_idx, column=1, value=label).font = Font(name="Calibri", bold=True)
            for c_idx, insurer in enumerate(columns_map.keys(), start=2):
                ws.cell(row=2, column=c_idx, value=insurer).font = Font(name="Calibri", bold=True)
                ws.cell(row=r_idx, column=c_idx, value=columns_map[insurer][r_idx-3])
                
        output_health = io.BytesIO()
        wb.save(output_health)
        
        st.download_button(
            label="📥 Download Health GMC QCR Matrix (Excel)",
            data=output_health.getvalue(),
            file_name="GMC_Comparison_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch"
        )

if __name__ == "__main__":
    main()
